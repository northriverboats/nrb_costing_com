#!/usr/bin/env python
# vim expandtab shiftwidth=2 softtabstop=2
"""
NRB COMMERCIAL COSTING SHEET GENERATOR

The files in the RESOURCES folder are labeled BOM but in reality is just a list
of parts from each department/motor/trailer that can go on a BOM

The files in the folder BOATS are both boat files and cabin files but are the
real BOM files

resources = parts that can be used on a bom
boms = a list of parts used in building a cabin or boat

DEBUG to log file
ERROR to log file and screen
CRITICAL to log file, screen and email
"""
import datetime
import logging
import logging.handlers
import os
import pprint
import sys
import traceback
from copy import deepcopy
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Union

import click
from dotenv import load_dotenv  # pylint: disable=import-error
import openpyxl  # pylint: disable=import-error

#
# ==================== Low Level Utilities
#
def resource_path(relative_path: Union[str, Path]) -> Path:
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # pylint: disable=protected-access
        base_path = Path(sys._MEIPASS)  # type: ignore
    except AttributeError:
        base_path = Path.cwd()

    return base_path / relative_path

env_path = resource_path('.env')
load_dotenv(dotenv_path=env_path)

DATABASE: Path = Path(os.environ.get('DATABASE', ''))
SHEETS_FOLDER: Path = Path(os.environ.get('SHEETS_FOLDER', ''))
BOATS_FOLDER: Path = Path(os.environ.get('BOATS_FOLDER', ''))
RESOURCES_FOLDER: Path = Path(os.environ.get('RESOURCES_FOLDER', ''))
TEMPLATE_FILE: Path = Path(os.environ.get('TEMPLATE_FILE', ''))
MASTER_FILE: Path = Path(os.environ.get('MASTER_FILE', ''))
CONSUMABLES_FILE: Path = Path(os.environ.get('CONSUMABLES_FILE', ''))
HOURLY_RATES_FILE: Path = Path(os.environ.get('HOURLY_RATES_FILE', ''))
MARK_UPS_FILE: Path = Path(os.environ.get('MARK_UPS_FILE', ''))
MAIL_SERVER: str = str(os.environ.get("MAIL_SERVER", ''))
MAIL_FROM: str = str(os.environ.get("MAIL_FROM", ''))
MAIL_TO: str = str(os.environ.get("MAIL_TO", ''))

_VERBOSE:  List[int] = [0]

#
# ==================== ENALBE LOGGING
# DEBUG + = to rotating log files in current directory
# INFO + = to stdout
# CRITICAL + = to email

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')

consoleHandler = logging.StreamHandler(sys.stdout)
consoleHandler.setLevel(logging.DEBUG)
consoleHandler.setFormatter(formatter)

fileHandler = logging.handlers.RotatingFileHandler(
    filename="error.log",
    maxBytes=1024000,
    backupCount=10,
    mode="a"
)
fileHandler.setLevel(logging.INFO)
fileHandler.setFormatter(formatter)

smtpHandler = logging.handlers.SMTPHandler(
    mailhost=MAIL_SERVER,
    fromaddr=MAIL_FROM,
    toaddrs=MAIL_TO,
    subject="alert!"
)
smtpHandler.setLevel(logging.CRITICAL)
smtpHandler.setFormatter(formatter)

logger.addHandler(consoleHandler)
logger.addHandler(fileHandler)
logger.addHandler(smtpHandler)


#
# ==================== Custom Errors
#
class NRBError(Exception):
    """Base class for all NRB errors"""
    def __init__(self, *args):
        if args:
            self.message = args[0]
        else:
            self.message = None
        super().__init__(self)

    def __str__(self):
        if self.message:
            return f"NRB Error, {self.message}"
        return 'NRB Error has been raised'


#
# ==================== Utility clases
#
def status_msg(msg: str, level: int, nl: bool = True) -> None:
    """output message if verbosity is sufficent"""
    if _VERBOSE[0] >= level:
        click.echo(msg, nl=nl)

# ==================== Dataclasses
#
@dataclass
class BoatModel:
    """Information on which sheets make a boat costing sheet"""
    sheet1: str
    sheet2: str
    folder: str

@dataclass(order=True)
class Resource:
    """BOM Part Information """
    # pylint: disable=too-many-instance-attributes
    # Eight is reasonable in this case.
    oempart: str
    description: str = field(compare=False)
    uom: str = field(compare=False)
    unitprice: float = field(compare=False)
    oem: str = field(compare=False)
    vendorpart: str = field(compare=False)
    vendor: str = field(compare=False)
    updated: datetime.datetime = field(compare=False)

@dataclass
class Consumable:
    """Consumables rate by department"""
    dept: str
    percent: float

@dataclass
class HourlyRate:
    """Hourly rate by department"""
    dept: str
    rate: float

@dataclass
class MarkUp:
    """Mark-up rates by deprtment"""
    policy: str
    markup_1: float
    markup_2: float
    discount: float

@dataclass(order=True)
class BomPart:
    """Part Information from Section of a BOM Parts Sheet"""
    part: Optional[str]
    qty: float = field(compare=False)
    smallest: Optional[float] = field(compare=False)
    biggest: Optional[float] = field(compare=False)
    percent: Optional[float] = field(compare=False)  # FT field

@dataclass(order=True)
class BomSection:
    """Group of BOM Parts"""
    name: str
    parts: List[BomPart] = field(compare=False)

@dataclass(order=True)
class Bom:
    """BOM sheet"""
    name: str
    smallest: float = field(compare=False)
    biggest: float = field(compare=False)
    sizes: List[float] = field(compare=False)
    sections: List[BomSection] = field(compare=False)

#
# ==================== Low Level Functions
#

def make_bom_part(row) -> BomPart:
    """Create bom part from row in spreadsheet"""
    qty: float = float(row[0].value)
    smallest: Optional[float] = (
        None if row[1].value is None else float(row[1].value))
    biggest: Optional[float] = (
        None if row[2].value is None else float(row[2].value))
    percent: Optional[float] = (
        None if row[3].value is None else float(row[3].value))
    part: Optional[str] = None if row[5].value is None else row[5].value
    return BomPart(part, qty, smallest, biggest, percent)

def section_add_part(section: BomSection, part: BomPart) -> None:
    """Insert new part or add qty to existing part"""
    found = [item for item in section.parts if item.part == part.part]
    if found:
        found[0].qty += part.qty
    else:
        section.parts.append(part)

def load_boat_models(master_file: Path) -> List[BoatModel]:
    """Build master list of sheets to combine to create costing sheets"""
    status_msg('Loading Boat Models', 1)
    status_msg(f'  {master_file.name}', 2)
    try:
        xlsx = openpyxl.load_workbook(master_file.as_posix(), data_only=True)
        sheet: openpyxl.worksheet.worksheet.Worksheet = xlsx.active
        boats: List[BoatModel] = []
        for row in sheet.iter_rows(min_row=2, max_col=3):
            if not isinstance(row[0].value, str):
                continue
            boat: BoatModel = BoatModel(
                row[0].value,
                row[1].value,
                row[2].value)
            boats.append(boat)
        xlsx.close()
    except (FileNotFoundError, PermissionError):
        boats: List[BoatModel] = []
    return boats

def load_resource_file(resource_file: Path) -> List[Resource]:
    """Read resource sheet"""
    status_msg(f'  {resource_file.name}', 2)
    try:
        xlsx = openpyxl.load_workbook(resource_file.as_posix(), data_only=True)
        sheet: openpyxl.worksheet.worksheet.Worksheet = xlsx.active
        resources: List[Resource] = []
        for row in sheet.iter_rows(min_row=2, max_col=8):
            if not isinstance(row[0].value, str):
                continue
            resource: Resource = Resource(
                row[0].value,
                row[1].value,
                row[2].value,
                float(row[3].value),
                row[4].value,
                row[5].value,
                row[6].value,
                row[7].value)
            resources.append(resource)
        xlsx.close()
    except (FileNotFoundError, PermissionError):
        pass
    return resources

def load_resources(resource_folder: Path) -> List[Resource]:
    """Load all resource files"""
    status_msg('Loading Resources', 1)
    resource_files: List[Path] = [
        sheet
        for sheet in find_excel_files_in_dir(resource_folder)
        if sheet.name.startswith('BOM ')]

    resources: List[Resource] = []
    for resource_file in resource_files:
        resources += load_resource_file(resource_file)
    return resources

def find_excel_files_in_dir(base: Union[str, Path]) -> List[Path]:
    """get list of spreadsheets in folder"""
    if isinstance(base, str):
        base = Path(base)
    return [sheets for sheets in base.glob('[!~]*.xlsx')]  # pylint: disable=unnecessary-comprehension

def load_consumables(resource_file: Path) -> List[Consumable]:
    """Read consuables sheet"""
    status_msg('Loading Consumables', 1)
    status_msg(f'  {resource_file.name}', 2)
    try:
        xlsx = openpyxl.load_workbook(resource_file.as_posix(), data_only=True)
        sheet: openpyxl.worksheet.worksheet.Worksheet = xlsx.active
        consumables: List[Consumable] = []
        for row in sheet.iter_rows(min_row=2, max_col=2):
            if not isinstance(row[0].value, str):
                continue
            consumable: Consumable = Consumable(
                row[0].value,
                float(row[1].value))
            consumables.append(consumable)
        xlsx.close()
    except (FileNotFoundError, PermissionError):
        pass
    return consumables

def load_hourly_rates(resource_file: Path) -> List[HourlyRate]:
    """Read hourly rates sheet"""
    status_msg('Loading Hourly Rates', 1)
    status_msg(f'  {resource_file.name}', 2)
    try:
        xlsx = openpyxl.load_workbook(resource_file.as_posix(), data_only=True)
        sheet: openpyxl.worksheet.worksheet.Worksheet = xlsx.active
        hourly_rates: List[HourlyRate] = []
        for row in sheet.iter_rows(min_row=2, max_col=2):
            if not isinstance(row[0].value, str):
                continue
            hourly_rate: HourlyRate = HourlyRate(
                row[0].value,
                float(row[1].value))
            hourly_rates.append(hourly_rate)
        xlsx.close()
    except (FileNotFoundError, PermissionError):
        pass
    return hourly_rates

def load_mark_ups(resource_file: Path) -> List[MarkUp]:
    """read makrkup file into object """
    status_msg('Loading Mark Ups', 1)
    status_msg(f'  {resource_file.name}', 2)
    try:
        xlsx = openpyxl.load_workbook(resource_file.as_posix(), data_only=True)
        sheet: openpyxl.worksheet.worksheet.Worksheet = xlsx.active
        mark_ups: List[MarkUp] = []
        for row in sheet.iter_rows(min_row=2, max_col=4):
            if not isinstance(row[0].value, str):
                continue
            mark_up: MarkUp = MarkUp(
                row[0].value,
                float(row[1].value),
                float(row[2].value),
                float(row[3].value))
            mark_ups.append(mark_up)
        xlsx.close()
    except (FileNotFoundError, PermissionError):
        pass
    return mark_ups


# ==================== Build BOM Functions
def get_hull_sizes(sheet: openpyxl.worksheet.worksheet.Worksheet) -> List:
    """find all hull sizes listed in sheet"""
    sizes = []
    for values  in sheet.iter_rows(
            min_row=1,max_row=1,min_col=13,values_only=True):
        for value in values:
            if value:
                sizes.append(float(value))
    return sizes

def get_bom_sections(sheet: openpyxl.worksheet.worksheet.Worksheet
                    ) -> List[BomSection]:
    """read in BOM items into sections"""
    sections: List[BomSection] = []
    for row in sheet.iter_rows(min_row=18,max_col=8):
        qty: Optional[Union[str, int, float]] = row[0].value
        if isinstance(qty, str) and qty != "QTY":
            if 'section' in locals():
                sections.append(section)
            section: BomSection = BomSection(qty, [])
        elif isinstance(qty, (float, int)):
            bom_part: BomPart = BomPart(
                str(row[5].value),
                float(row[0].value),
                float(0 if row[1].value is None else row[1].value),
                float(0 if row[2].value is None else row[2].value),
                float(0 if row[3].value is None else row[3].value))
            section.parts.append(bom_part)
    sections.append(section)
    return sections


def load_bom(bom_file: Path) -> Bom:
    """load individual BOM sheet"""
    status_msg(f'  {bom_file.name}', 2)
    try:
        xlsx = openpyxl.load_workbook(bom_file.as_posix(), data_only=True)
        sheet: openpyxl.worksheet.worksheet.Worksheet = xlsx.active

        name: str = str(sheet["A1"].value)
        smallest: float = float(
            0 if sheet["M1"].value == "ANY" else sheet["G13"].value)
        biggest:float  = float(
            0 if sheet["M1"].value == "ANY" else sheet["G14"].value)
        sizes = [] if smallest == 0 else get_hull_sizes(sheet)
        resources: List[BomSection] = get_bom_sections(sheet)
        bom: Bom = Bom(name, smallest, biggest, sizes, resources)
        xlsx.close()
    except (FileNotFoundError, PermissionError):
        pass
    return bom

def load_boms(bom_folder: Path) -> List[Bom]:
    """load all BOM sheets"""
    status_msg('Loading BOMs', 1)
    bom_files: List[Path] = find_excel_files_in_dir(bom_folder)
    boms: List[Bom] = []
    for bom_file in bom_files:
        bom = load_bom(bom_file)
        boms.append(bom)
    return boms

#
# ==================== Merge BOMs
#
def bom_merge_section(parts1: List[BomPart], parts2: List[BomPart]) -> None:
    """Merege two sections"""
    for part2 in parts2:
        part1 = next(
            iter([part for part in parts1 if part.part == part2.part]),None)
        if part1:
            part1.qty += part2.qty
        else:
            parts1.append(deepcopy(part2))

def bom_merge(bom1: Bom, bom2: Bom) -> Bom:
    """Merge two BOMs creating a new BOM in the process"""
    bom: Bom = deepcopy(bom1)
    for section1, section2 in zip(bom.sections, bom2.sections):
        bom_merge_section(section1.parts, section2.parts)
    return bom

def get_bom(boms: List[Bom], model: BoatModel) -> Bom:
    """Combine sheets if necessary and return BOM
       Assumes if sheet is not None that there will be a match"""
    bom1: Bom = next(iter([bom for bom in boms if bom.name == model.sheet1]))
    bom2: Bom = Bom('', 0.0, 0.0, [], []) if model.sheet2 else next(
        iter([bom for bom in boms if bom.name == model.sheet2]))
    return bom_merge(bom1, bom2)


#
# ==================== Main Entry Point
#
@click.command()
@click.option('-v', '--verbose', count=True)
def main(verbose: int) -> None:
    """ main program entry point """
    _VERBOSE[0] = verbose
    try:
        models: List[BoatModel] = load_boat_models(MASTER_FILE)
        resources: List[Resource] = load_resources(RESOURCES_FOLDER)
        consumables: List[Consumable] = load_consumables(CONSUMABLES_FILE)
        hourly_rates: List[HourlyRate] = load_hourly_rates(HOURLY_RATES_FILE)
        mark_ups: List[MarkUp] = load_mark_ups(MARK_UPS_FILE)
        boms: List[Bom] = load_boms(BOATS_FOLDER)

        status_msg(f'Models: {len(models)}   ', 1, nl=False)
        status_msg(f'Resources: {len(resources)}   ', 1, nl=False)
        status_msg(f'Consumables: {len(consumables)}   ', 1,  nl=False)
        status_msg(f'Hourly Rates: {len(hourly_rates)}   ', 1, nl=False)
        status_msg(f'Mark Ups: {len(mark_ups)}   ', 1, nl=False)
        status_msg(f'BOMs: {len(boms)}   ', 1)
        # click.echo(pprint.pformat(resources, width=210))
        status_msg(pprint.pformat(models, width=140), 3)
    except Exception:
        logger.critical(traceback.format_exc())
        raise
    finally:
        # program terminates normally
        sys.exit()

if __name__ == "__main__":
    main()  # pylint: disable=E1120
