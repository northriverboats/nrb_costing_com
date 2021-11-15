#!/usr/bin/env python
# vim expandtab shiftwidth=2 softtabstop=2
"""
Load model data from sheet

Pass in master_file return data structure
"""
from dataclasses import dataclass
from pathlib import Path
from openpyxl import load_workbook # pylint: disable=import-error
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from utility import status_msg


@dataclass
class Consumable:
    """Consumables rate by department"""
    dept: str
    percent: float


def load_consumables(xlsx_file: Path) -> dict[str, Consumable]:
    """Read consuables sheet"""
    status_msg('Loading Consumables', 1)
    status_msg(f'  {xlsx_file.name}', 2)
    xlsx: Workbook = load_workbook(xlsx_file.as_posix(), data_only=True)
    sheet: Worksheet = xlsx.active
    consumables: dict[str, Consumable] = {}
    for row in sheet.iter_rows(min_row=2, max_col=2):
        if not isinstance(row[0].value, str):
            continue
        consumable: Consumable = Consumable(
            row[0].value,
            float(row[1].value))
        status_msg(f"    {consumable}", 3)
        consumables[row[0].value] = consumable
    xlsx.close()
    return consumables

if __name__ == "__main__":
    pass
