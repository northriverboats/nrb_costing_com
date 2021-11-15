#!/usr/bin/env python
# vim expandtab shiftwidth=2 softtabstop=2
"""
Load model data from sheet

Pass in master_file return data structure
"""
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Union
from openpyxl import load_workbook # pylint: disable=import-error
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from utility import status_msg

@dataclass(order=True)
class BomPart:
    """Part Information from Section of a BOM Parts Sheet"""
    part: str
    qty: float = field(compare=False)
    smallest: Optional[float] = field(compare=False)
    biggest: Optional[float] = field(compare=False)
    percent: Optional[float] = field(compare=False)  # FT field

@dataclass(order=True)
class BomSection:
    """Group of BOM Parts"""
    name: str
    parts: dict[str, BomPart] = field(compare=False)

@dataclass(order=True)
class Bom:
    """BOM sheet"""
    name: str
    beam: str = field(compare=False)
    smallest: float = field(compare=False)
    biggest: float = field(compare=False)
    sizes: list[float] = field(compare=False)
    sections: list[BomSection] = field(compare=False)


def find_excel_files_in_dir(base: Path) -> list[Path]:
    """get list of spreadsheets in folder"""
    return list(base.glob('[!~]*.xlsx'))

def make_bom_part(row) -> BomPart:
    """Create bom part from row in spreadsheet"""
    qty: float = float(row[0].value)
    smallest: Optional[float] = (
        None if row[1].value is None else float(row[1].value))
    biggest: Optional[float] = (
        None if row[2].value is None else float(row[2].value))
    percent: Optional[float] = (
        None if row[3].value is None else float(row[3].value))
    part: str = str(row[5].value)
    return BomPart(part, qty, smallest, biggest, percent)

def section_add_part(parts: dict[str, BomPart], part: BomPart) -> None:
    """Insert new part or add qty to existing part"""
    try:
        item: BomPart = parts[part.part]
        item.qty += part.qty
    except KeyError:
        parts[part.part] = part


def get_hull_sizes(sheet: Worksheet) -> list:
    """find all hull sizes listed in sheet"""
    sizes = []
    for values  in sheet.iter_rows(
            min_row=1,max_row=1,min_col=13,values_only=True):
        for value in values:
            if value:
                sizes.append(float(value))
    return sizes

def get_bom_sections(sheet: Worksheet) -> list[BomSection]:
    """read in BOM items into sections"""
    sections: list[BomSection] = []
    for row in sheet.iter_rows(min_row=18,max_col=8):
        qty: Optional[Union[str, int, float]] = row[0].value
        if isinstance(qty, str) and qty != "QTY":
            if qty == 'CANVAS':
                continue
            if 'section' in locals():
                sections.append(section)
            section: BomSection = BomSection(qty, {})
        elif isinstance(qty, (float, int)):
            bom_part: BomPart = BomPart(
                str(row[5].value),
                float(row[0].value),
                float(0 if row[1].value is None else row[1].value),
                float(0 if row[2].value is None else row[2].value),
                float(0 if row[3].value is None else row[3].value))
            section.parts[str(row[5].value)] = bom_part
    sections.append(section)
    return sections

def load_bom(xlsx_file: Path) -> Bom:
    """load individual BOM sheet"""
    status_msg(f'  {xlsx_file.name}', 2)
    xlsx: Workbook = load_workbook(xlsx_file.as_posix(), data_only=True)
    sheet: Worksheet = xlsx.active

    name: str = str(sheet["A1"].value)
    beam: str = "" if sheet["A2"].value is None else sheet["A2"].value
    smallest: float = float(
        0 if sheet["M1"].value == "ANY" else sheet["G13"].value)
    biggest:float  = float(
        0 if sheet["M1"].value == "ANY" else sheet["G14"].value)
    sizes = [] if smallest == 0 else get_hull_sizes(sheet)
    resources: list[BomSection] = get_bom_sections(sheet)
    bom: Bom = Bom(name, beam, smallest, biggest, sizes, resources)
    xlsx.close()
    return bom

def load_boms(bom_folder: Path) -> list[Bom]:
    """load all BOM sheets"""
    status_msg('Loading BOMs', 1)
    bom_files: list[Path] = find_excel_files_in_dir(bom_folder)
    boms: list[Bom] = []
    for bom_file in bom_files:
        bom = load_bom(bom_file)
        boms.append(bom)
    return boms

if __name__ == "__main__":
    pass
