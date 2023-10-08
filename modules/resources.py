#!/usr/bin/env python
# vim expandtab shiftwidth=2 softtabstop=2
"""
Load resource data from sheet

If I1 has "Dealer Net Price" then there is net pricing to capture
"""
from datetime import datetime
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
from dataclasses_json import DataClassJsonMixin
from openpyxl import load_workbook # pylint: disable=import-error
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from .utilities import status_msg


@dataclass(order=True)
class Resource(DataClassJsonMixin):
    """BOM Part Information """
    # pylint: disable=too-many-instance-attributes
    # Eight is reasonable in this case.
    oempart: str
    description: str = field(compare=False)
    uom: str = field(compare=False)
    unitprice: float = field(compare=False)
    oem: str = field(compare=False)
    vendorpart: str = field(compare=False)
    vendor: str = field(compare=False)
    updated: datetime = field(compare=False)
    dealer_net: float = field(compare=False)

@dataclass
class Resources(DataClassJsonMixin):
    """Dictionary of all resources"""
    resources: dict[str, Resource]

def find_excel_files_in_dir(base: Path) -> list[Path]:
    """get list of spreadsheets in folder"""
    return list(base.glob('[!~]*.xlsx'))

def load_resource_file(xlsx_file: Path) -> Resources:
    """Read resource sheet"""
    status_msg(f'  {xlsx_file.name}', 2)
    xlsx: Workbook = load_workbook(xlsx_file.as_posix(), data_only=True)
    sheet: Worksheet = xlsx.active
    net_price: bool = sheet['I1'].value == "Dealer Net Price"
    all_resources: Resources = Resources({})
    for row in sheet.iter_rows(min_row=2, max_col=9):
        if not isinstance(row[0].value, str):
            continue
        net_value = 0.0
        if net_price:
            net_value = float(row[8].value)
        resource: Resource = Resource(
            row[0].value,
            row[1].value,
            row[2].value,
            float(row[3].value),
            row[4].value,
            row[5].value,
            row[6].value,
            row[7].value,
            net_value)
        all_resources.resources[row[0].value] = resource
        status_msg(f"    {resource}",3)
    xlsx.close()
    return all_resources

def load_resources(resource_folder: Path) -> Resources:
    """Load all resource files"""
    status_msg('Loading Resources', 1)
    resource_files: list[Path] = [
        sheet
        for sheet in find_excel_files_in_dir(resource_folder)
        if sheet.name.startswith('BOM ')]

    all_resources: Resources = Resources({})
    for resource_file in resource_files:
        if resource_file == "K:\2024 HGAC - TEST\RESOURCE\BOM OUTBOARD MOTORS.xlsx":
            _ = 1 + 1
            pass
            _ = 1 + 1
        all_resources.resources.update(
            load_resource_file(resource_file).resources)
    return all_resources

if __name__ == "__main__":
    pass
