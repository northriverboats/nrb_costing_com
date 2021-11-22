#!/usr/bin/env python
# vim expandtab shiftwidth=4 softtabstop=4
"""
Generate Costing Sheets
"""
from copy import copy, deepcopy
# from datetime import datetime
from dataclasses import dataclass, field
from typing import TypedDict
from pathlib import Path
import os
import re
from openpyxl import load_workbook # pylint: disable=import-error
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from .boms import Bom, BomPart, BomSection
from .models import Model
from .utilities import (logger, normalize_size, status_msg, SHEETS_FOLDER,
                        TEMPLATE_FILE,)

@dataclass
class SheetRange:
    """Range for caclating formula offsets"""
    name: str   # name of range
    start: int  # start area for computing offsets
    first: int  # first row with formulas in it
    end: int  # end aera for computing offsets/last row with formulats in it
    subtotal: int # where subotal is for section
    max_delete: int # max number of rows that can be deleted
    offset: int = field(init=False) # relative offset
    lines: int = field(init=False) # number of formula lines
    add_del: int = field(init=False) # number of lines to add or delete
    parts: int = field(init=False) # number of parts in section

    def __post_init__(self):
        """initialize range"""
        self.offset = 0
        self.parts = 0
        self.add_del = 0
        self.lines = self.end + 1 - self.first

@dataclass
class SheetRanges:
    """List for computing offect for formulas"""
    ranges: list[SheetRange] = field(default_factory=list)

    def offset(self, reference: str) -> str:
        """update offset for cell reference
        * does not handle links
        * must list functions by name
        * will lnot handle named ranges
        * parse Sheet ! references but will update offsets though they are not
          on the curernt sheet

        Arguments:
            reference: str -- cell reference such as D22

        Returns
            str -- cell reference such as D14
        """
        blob: list = re.findall(r"(.+!\$?[A-Z]+\$?|\$?[A-Z]+\$?)(\d+)",
                          reference)
        if not blob:
            return reference
        col: str = blob[0][0]
        row: int = int(blob[0][1])
        for rows in self.ranges:
            if rows.start <= row <= rows.end:
                return col + str(row + rows.offset)
        return reference

    def adjust(self, parts_tally: list[int]) -> None:
        """adjust row offsets
        Arguments:
            parts_tally: list[int] -- how many parts each section has
        """
        offset: int = 0
        for section, tally in zip(self.ranges, parts_tally):
            if tally - section.lines < 0:
                section.add_del = -(min(section.lines - tally,
                                        section.max_delete))
            else:
                section.add_del = tally - section.lines
            section.parts = tally
            section.offset = offset
            section.first += offset
            offset += section.add_del
            section.end += offset
            section.subtotal += offset
            section.lines = section.end + 1 - section.first

    def offset_formula(self, formula: str) -> str:
        """compute formula offsets-- will not handle named ranges

        Arguements:
          formula: str -- formula in

        Return:
          str -- formula out
        """
        if formula[0] != "=":
            return formula
        result: str = "="
        reference: str = ""
        for char in formula[1:]:
            if reference in ["SUM", "VLOOKUP", "HLOOKUP"]:
                result += reference + char
                reference = ""
                continue

            if char not in "(*/+-):":
                reference += char
                continue
            if reference:
                result += self.offset(reference)
                reference = ""
            result += char
        if reference:
            result += self.offset(reference)
            reference = ""
        return result

    def find(self, name: str) -> SheetRange:
        """find matching sheet range"""
        match: list[SheetRange] = [
            offset for offset in self.ranges if offset.name == name]
        return match[0]

    def show(self):
        """show range table"""
        for rows in self.ranges:
            print(rows)


ranges = SheetRanges()
ranges.ranges.append(SheetRange('FABRICATION', 1, 14, 17, 20, 1))
ranges.ranges.append(SheetRange('PAINT', 18, 26, 38, 40, 9))
ranges.ranges.append(SheetRange('UNUSED', 39, 49, 70, 72, 0))
ranges.ranges.append(SheetRange('OUTFITTING',71, 77, 139, 141, 68))
ranges.ranges.append(SheetRange('BIG TICKET ITEMS', 140, 148, 149, 151, 0))
ranges.ranges.append(SheetRange('OUTBOARD MOTORS', 150, 156, 158, 160, 0))
ranges.ranges.append(SheetRange('INBOARD MOTORS & JETS', 159, 165, 168, 170,
                                0))
ranges.ranges.append(SheetRange('TRAILER', 169, 175, 175, 177, 0))
ranges.ranges.append(SheetRange('TOTALS', 176, 177, 235, 235, 0))



# UTILITY FUNCTIONS ===========================================================
class FileNameInfo(TypedDict):
    """file name info"""
    size: str
    model: str
    option: str
    with_options: str
    size_with_options: str
    file_name: Path

def build_name(size: float, model: Model, folder: str) -> FileNameInfo:
    """build file name for sheet

    Arguments:
        size: float  -- length of boat ending in .0 or .5
        model: Model -- name of model of boat

    Returns:
        dict -- size of boat as text  such as 18' or 18'6"
                model: of boat as text
                option: of boat as text
                with_options: text of model + option
                size_with_options: size + model + option
                file_name: full absolute path to file
    """
    option: str = "" if model.sheet2 is None else ' ' + model.sheet2
    size_with_options: str = normalize_size(size) + ' ' + model.sheet1 + option
    file_name: Path = SHEETS_FOLDER / folder / (size_with_options + '.xlsx')
    name: FileNameInfo = {
        'size': normalize_size(size),
        'model': model.sheet1,
        'option': option,
        'with_options': model.sheet1 + option,
        'size_with_options': size_with_options,
        'file_name': file_name,
    }
    return name


# BOM MANIPULATION FUNCTIONS ==================================================
def ordered_parts(parts: dict[str, BomPart], name: str) -> list[str]:
    """set correct sort order for parts for each section. Outftting is sorted
    by vender then part number. All other sections are ordered by part number

    Arguments:
        parts: dict[str, BomPart] -- dictionary of parts with the key being a
                                     part number
        name: str -- name of section

    Returns:
        list[str] -- Correctly ordered list of keys
    """
    if name != 'OUTFITTING':
        return sorted(parts)
    return sorted(parts, key=lambda k: (parts[k].vendor, k))


def bom_merge_section(target_parts: dict[str, BomPart],
                      source_parts: dict[str, BomPart]) -> None:
    """Merege two sections

    Arguments:
        target_parts: dict[str, BomPart] -- longer dict of BomParts to append
                                            to or merge part with
        source_parts: dict[str, BomPart] -- shorter dict of BomParts to add or
                                            merge parts from

    Returns:
        None -- target_parts modified in place
    """
    for key in source_parts:
        try:
            target_part: BomPart = target_parts[key]
            target_part.qty += source_parts[key].qty
        except KeyError:
            target_parts[key] = deepcopy(source_parts[key])

def bom_merge(target_bom: Bom, source_bom: Bom) -> Bom:
    """Merge two BOMs creating a new BOM in the process. Does not modify either
    original Bom.

    Arguments:
        target_bom: Bom -- sheet1 bom we want to merge parts into
        source_bom: Bom -- sheet2 bom with parts to be added/merged

    Returns:
        Bom -- new merged list of both boms combined
    """
    bom: Bom = deepcopy(target_bom)
    for target, source in zip(bom.sections, source_bom.sections):
        bom_merge_section(target.parts, source.parts)
    return bom

def get_bom(boms: dict[str, Bom], model: Model) -> Bom:
    """Merges sheets if necessary and returns a BOM.
    Assumes if sheet is not None that there will be a match

    Arguments:
        bom: list[Bom] --
        model: Model -- sheet1 can not be None and must be found
                        sheet2 can be None but *must* be found if not None

    Returns:
        Bom -- Returns new Bom of combined Bom(s)
    """
    bom1: Bom = (boms[model.sheet1]
                 if model.sheet1 in boms
                 else  Bom('', "", 0.0, 0.0, [], []))
    bom2: Bom = (boms[model.sheet2]
                 if model.sheet2 in boms
                 else  Bom('', "", 0.0, 0.0, [], []))
    if bom1.name == "":
        logger.debug("bom1 not found error %s", model.sheet1)
    if bom2.name == "":
        logger.debug("bom2 not found error %s %s  %s",
                     model.sheet1, model.sheet2, model.folder)
    return bom_merge(bom1, bom2)


def compute_section_sizes(bom_sections: list[BomSection],) -> SheetRanges:
    """compute size of sections and appropriate offsets
    * insert blank sections for UNUSED and TOTALS so that list[BomSections] and

    Arguements:
        bom_sections: list[BomSection] -- parts sections of BOM

    Returns:
        SheetRanges -- SheetRanges with oupdated section sizes
    """
    # list[BomSection] .name: str .parts: dict[BomPart]
    parts_tally: list[int] = [
        len(section.parts)
        for section in bom_sections]
    parts_tally = parts_tally[:2] + [0] + parts_tally[2:] + [0]
    offsets: SheetRanges = deepcopy(ranges)
    pprint(offsets.ranges)
    offsets.adjust(parts_tally)
    return offsets


def filter_bom(original_bom: Bom, size: float) -> Bom:
    """fitler out parts based on size if necessary
    also correcting costs as necessary

    Arguments:
        bom: Bom -- Bom with parts sections
        size: float -- size of boat to filter for

    Returns:
        Bom -- Deepcopy Bom with correct parts
    """
    bom: Bom = deepcopy(original_bom)
    for section in bom.sections:
        section.parts = {
            k:v
            for k, v in section.parts.items()
            if (v.smallest == 0 or size >= v.smallest) and
            (v.biggest == 0 or size <= v.biggest)
        }
        for v in section.parts.values():
            if v.percent:
                v.unitprice = size / v.percent * v.unitprice
    return bom


# WRITING SHEET FUNCTIONS =====================================================
def resize_sections(section_sizes: SheetRanges, sheet: Worksheet) -> None:
    """resize sections of sheet by removig rows or adding rows and formatting
    them. Will also redo-the formula references in the range and subtotal

    Arguments:
        section_sizes: SheetRanges -- data on how to resize the sections
        sheet: Worksheet -- handle to xlsx worksheet

    Returns:
        None
    """
    for section_size in section_sizes.ranges:
        print(f"{section_size.name}  -- {section_size.add_del}")
    print(sheet)


def generate_sections(bom: Bom,
                      section_sizes: SheetRanges,
                      sheet: Worksheet) -> None:
    """Manage filling in sections

    Works by iterating over the sections of the sheet from the last section to
    the first section. Order matters here.
    * bom.Sections were built from first section to last section. Needs to be
      iterated over in reverse order
    * sheet_ranges was constructed in reverse order

    Arguements:
    """
    for section in bom.sections:
        offset = section_sizes.find(section.name)
        # generate_section(sheet, section, size, info)
        print(f"{section.name:24} {offset}")
    print()

def generate_heading(bom: Bom,
                     file_name_info: FileNameInfo,
                     sheet: Worksheet) -> None:
    """Fill out heading at top of sheet
    Arguments:
        bom: Bom -- bom with information for all sizes of current model/option
        name: dict -- parts and full name of current sheet

    Returns:
        None
    """
    sheet["C4"].value = file_name_info['with_options'].title()
    sheet["C5"].value = file_name_info['size']
    sheet["C6"].value = bom.beam

def generate_sheet(filtered_bom: Bom, file_name_info: FileNameInfo) -> None:
    """genereate costing sheet

    Arguments:
        filterd_bom: Bom -- bom with only parts fr the current size
        name: dict -- parts and full name of current sheet
        file_name: Path -- filename with full pathing to xls sheet to be
                           created

    Returns:
        None
    """
    # create parent folder if necessay
    file_name_info['file_name'].parent.mkdir(parents=True, exist_ok=True)
    # caculate the size of each section
    section_sizes: SheetRanges = compute_section_sizes(filtered_bom.sections)

    # open templeate file
    xlsx: Workbook = load_workbook(TEMPLATE_FILE.as_posix(), data_only=False)
    sheet: Worksheet = xlsx.active

    # resize_sections(section_sizes, sheet)
    # generate_heading(filtered_bom, file_name_info, sheet)
    # generate_sections(filtered_bom, section_sizes, sheet)
    xlsx.save(os.path.abspath(str(file_name_info['file_name'])))


# MODEL/SIZE IETERATION FUNCTIONS =============================================
def generate_sheets_for_model(model: Model, bom: Bom) -> None:
    """"cycle through each size to create sheets
    * build the filname and size as as a text name
    * filter out parts that are not needed for this size of boat and correct
      costing for size of boat if necessay
    * computing section sizes is done in genereate_sheet

    Arguments:
        model: Model -- Model of boat to process
        bom: Bom -- Base Bom for model

    Returns:
        None
    """
    status_msg(f"  {model.folder}", 1)
    for size in bom.sizes:
        file_name_info: FileNameInfo = build_name(size, model, model.folder)
        status_msg(f"    {file_name_info['file_name']}", 2)
        filtered_bom = filter_bom(bom, size)
        generate_sheet(filtered_bom, file_name_info)


def junk():
    """Temp place for insert and format cells in sheet"""
    xlsx: Workbook = load_workbook(TEMPLATE_FILE.as_posix(), data_only=False)
    sheet: Worksheet = xlsx.active
    sheet.insert_rows(78,5)

    for col in range(1,16):
        cell = sheet.cell(row=77, column=col)
        style = copy(cell.style)
        font = copy(cell.font)
        border = copy(cell.border)
        number_format = copy(cell.number_format)
        alignment = copy(cell.alignment)

        for row in range(78,83):
            cell = sheet.cell(row=row, column=col)
            cell.style = copy(style)
            cell.font = copy(font)
            cell.border = copy(border)
            cell.number_format = copy(number_format)
            cell.alignment = copy(alignment)
            if col == 5:
                cell.value = 'ea'
            if col == 7:
                cell.value = f'=D{row}*F{row}'
            if col == 8:
                cell.value = 0
            if col == 9:
                cell.value = f'=G{row}+H{row}'

    xlsx.save(os.path.abspath(str('/home/fwarren/test.xlsx')))


def generate_sheets_for_all_models(models: dict[str, Model],
                                   boms: dict[str, Bom]) -> None:
    """" cycle through each sheet/option combo to create sheets

    Arguments:
        target_parts: dict[str, BomPart] -- longer dict of BomParts to append
                                            to or merge part with
        source_parts: dict[str, BomPart] -- shorter dict of BomParts to add or
                                            merge parts from

    Returns:
        None
    """
    status_msg("Merging", 1)
    # for model in models:  # fww
    for key in {"SOUNDER 8'6'' OPEN": models["SOUNDER 8'6'' OPEN"]}:  # fww
        # merge "model" sheet 1 Bom with "option" sheet 2 Bom
        bom: Bom = get_bom(boms, models[key])
        generate_sheets_for_model(models[key], bom)
        # for section in bom.sections:
        #   print(f"{section.name:25} {len(section.parts)}")

if __name__ == "__main__":
    pass
