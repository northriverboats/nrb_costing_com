#!/usr/bin/env python
# vim expandtab shiftwidth=2 softtabstop=2
"""
Load model data from sheet

Pass in master_file return data structure
"""
from dataclasses import dataclass
from dataclasses_json import DataClassJsonMixin
from pathlib import Path
from openpyxl import load_workbook # pylint: disable=import-error
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from .utilities import status_msg

@dataclass
class Consumable(DataClassJsonMixin):
    """Consumables rate"""
    rate: float

@dataclass
class Consumables(DataClassJsonMixin):
    """Consumables rate by department"""
    consumables: dict[str, Consumable]

def load_consumables(xlsx_file: Path) -> Consumables:
    """Read consuables sheet"""
    status_msg('Loading Consumables', 1)
    status_msg(f'  {xlsx_file.name}', 2)

    consumable: Consumable
    consumables: Consumables = Consumables({})
    
    xlsx: Workbook = load_workbook(xlsx_file.as_posix(), data_only=True)
    sheet: Worksheet = xlsx.active

    for row in sheet.iter_rows(min_row=1, max_col=2):
        dept, rate = [row[0].value, row[1].value]
        if not isinstance(dept, str):
            continue
        consumables.consumables[dept] = Consumable(float(rate))
        status_msg(f"    {dept:12.12}  {rate * 100:5.2f}%", 3)
    xlsx.close()
    return consumables

if __name__ == "__main__":
    pass
