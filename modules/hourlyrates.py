#!/usr/bin/env python
# vim expandtab shiftwidth=2 softtabstop=2
"""
Load model data from sheet

Pass in master_file return data structure
"""
from dataclasses import dataclass
from pathlib import Path
from dataclasses_json import DataClassJsonMixin
from openpyxl import load_workbook # pylint: disable=import-error
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from .utilities import status_msg


@dataclass
class HourlyRate(DataClassJsonMixin):
    """Hourly rate by department"""
    rate: float

@dataclass(order=True)
class HourlyRates(DataClassJsonMixin):
    """Hourly Rates"""
    hourly_rates: dict[str, HourlyRate]

def load_hourly_rates(xlsx_file: Path) -> HourlyRates:
    """Read hourly rates sheet"""
    status_msg('Loading Hourly Rates', 1)
    status_msg(f'  {xlsx_file.name}', 2)
    xlsx: Workbook = load_workbook(xlsx_file.as_posix(), data_only=True)
    sheet: Worksheet = xlsx.active
    hourly_rate: HourlyRate
    hourly_rates: HourlyRates = HourlyRates({})
    for row in sheet.iter_rows(min_row=1, max_col=2):
        if not isinstance(row[0].value, str):
            continue
        value: float = float(row[1].value)
        name: str = row[0].value
        hourly_rate = HourlyRate(float(value))
        status_msg(f"    {hourly_rate}", 3)
        hourly_rates.hourly_rates[name] = hourly_rate
    xlsx.close()
    return hourly_rates

if __name__ == "__main__":
    pass
