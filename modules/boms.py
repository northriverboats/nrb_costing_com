#!/usr/bin/env python
# vim expandtab shiftwidth=2 softtabstop=2
"""
Load model data from sheet

Pass in master_file return data structure
"""
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional, Union
from dataclasses_json import DataClassJsonMixin
from openpyxl import load_workbook # pylint: disable=import-error
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from .resources import Resource
from .utilities import status_msg

@dataclass(order=True)
class BomPart(DataClassJsonMixin):
    """Part Information from Section of a BOM Parts Sheet"""
    # pylint: disable=too-many-instance-attributes
    part: str
    qty: float = field(compare=False)
    smallest: float = field(compare=False)
    biggest: float = field(compare=False)
    percent: float = field(compare=False)  # FT field
    description: str = field(compare=False)
    uom: str = field(compare=False)
    unitprice: float = field(compare=False)
    oem: str = field(compare=False)
    vendorpart: str = field(compare=False)
    vendor: str = field(compare=False)
    updated: datetime = field(compare=False)

@dataclass(order=True)
class BomSection(DataClassJsonMixin):
    """Group of BOM Parts"""
    name: str
    parts: dict[str, list[BomPart]] = field(compare=False)

@dataclass(order=True)
class Bom(DataClassJsonMixin):
    """BOM sheet"""
    name: str
    beam: str = field(compare=False)
    smallest: float = field(compare=False)
    biggest: float = field(compare=False)
    sizes: dict[str, dict[str, float]] = field(compare=False)
    sections: list[BomSection] = field(compare=False)

@dataclass(order=True)
class Boms(DataClassJsonMixin):
    """BOM sheets"""
    boms: dict[str, Bom]

@dataclass()
class MergedPart(DataClassJsonMixin):
    """Part Information from Section of a BOM Parts Sheet"""
    # pylint: disable=too-many-instance-attributes
    qty: float
    description: str
    uom: str
    unitprice: float
    vendor: str
    updated: datetime
    total: float

# Output BOM type
@dataclass(order=True)
class MergedSection(DataClassJsonMixin):
    """Group of BOM Parts"""
    name: str
    parts: dict[str, MergedPart] = field(compare=False)
    total:float = field(compare=False, default=0)

@dataclass(order=True)
class MergedBom(DataClassJsonMixin):
    """BOM sheet"""
    name: str
    beam: str = field(compare=False)
    size: str = field(compare=False)
    labor: dict[str, float] = field(compare=False)
    sections: list[MergedSection] = field(compare=False)

HOURTYPES = {
    'Design Hours': 'Design / Drafting',
    'Fabrication Hours': 'Fabrication',
    'Paint Hours': 'Paint',
    'Outfitting Hours': 'Outfitting',
    'Canvas Hours': 'Canvas',
}

def find_excel_files_in_dir(base: Path) -> list[Path]:
    """get list of spreadsheets in folder"""
    return list(base.glob('[!~]*.xlsx'))

def make_bom_part(row, resources: dict[str, Resource]) -> BomPart:
    """Create bom part from row in spreadsheet"""
    qty: float = float(row[0].value)
    smallest: float = (
        0.0
        if row[1].value is None or row[1].value == ''
        else float(row[1].value))
    biggest: float = (
        0.0
        if row[2].value is None or row[2].value == ''
        else float(row[2].value))
    percent: float = (
        0.0
        if row[3].value is None or row[3].value == ''
        else float(row[3].value))
    part: str = str(row[5].value)

    # fww resolve at later point, we need to throw errors on fail
    if part in resources:
        resource = resources[part]
    else:
        resource = Resource(part, "Unknown", "EA", 0.0, "Unknown", part,
                            "Unknown", datetime(1999,12,31))
    return BomPart(part,
                   qty,
                   smallest,
                   biggest,
                   percent,
                   resource.description,
                   resource.uom,
                   resource.unitprice,
                   resource.oem,
                   resource.vendorpart,
                   resource.vendor,
                   resource.updated)

def section_add_part(parts: dict[str, list[BomPart]], part: BomPart) -> None:
    """Insert new part or add qty to existing part"""
    if part.part not in parts:
        parts[part.part] = []
    parts[part.part].append(part)


def get_hull_sizes(sheet: Worksheet) -> dict[str, dict[str, float]]:
    """find all hull sizes listed in sheet"""
    sizes: dict[str, dict[str, float]] = {}
    for values  in sheet.iter_rows(
            min_row=1,max_row=1,min_col=13,values_only=True):
        for value in values:
            if value:
                sizes[str(value)] = {}
    return sizes

def get_bom_sections(sheet: Worksheet,
                     resources: dict[str, Resource]) -> list[BomSection]:
    """read in BOM items into sections"""
    sections: list[BomSection] = []
    for row in sheet.iter_rows(min_row=18,max_col=8):
        qty: Optional[Union[str, int, float]] = row[0].value
        if isinstance(qty, str) and qty != "QTY":
            if qty == 'CANVAS':
                continue
            if 'section' in locals():
                sections.append(section)
            section: BomSection = BomSection(qty, {})
        elif isinstance(qty, (float, int)):
            bom_part: BomPart = make_bom_part(row, resources)
            status_msg(f"    {bom_part}",3)
            section_add_part(section.parts, bom_part)
    sections.append(section)
    return sections

def bom_hours(sheet: Worksheet,
              sizes: dict[str, dict[str, float]]) -> None:
    """read in labor  hours and update sizes array"""
    hours: Optional[Union[str, int, float]]
    maximum = len(sizes) * 4 + 11
    for row in sheet.iter_rows(min_row=14,max_col=maximum):
        name: Optional[str] = row[6].value
        if not (isinstance(name, str) and "Hours" in (name or '')):
            continue
        for index, size in enumerate(sizes):
            hours = row[14 + index * 4].value
            if not isinstance(hours, (int, float)):
                hours = 0.0
            sizes[size][ HOURTYPES[(name or '')] ] = hours

def load_bom(xlsx_file: Path, resources: dict[str, Resource]) -> Bom:
    """load individual BOM sheet"""
    status_msg(f'  {xlsx_file.name}', 2)
    xlsx: Workbook = load_workbook(xlsx_file.as_posix(), data_only=True)
    sheet: Worksheet = xlsx.active

    name: str = str(sheet["A1"].value)
    beam: str = "" if sheet["A2"].value is None else sheet["A2"].value
    smallest: float = float(
        0 if sheet["M1"].value == "ANY" else sheet["G13"].value)
    biggest:float  = float(
        0 if sheet["M1"].value == "ANY" else sheet["G14"].value)
    sizes = {"0": {}} if smallest == 0 else get_hull_sizes(sheet)
    sections: list[BomSection] = get_bom_sections(sheet, resources)
    bom_hours(sheet, sizes)
    bom: Bom = Bom(name, beam, smallest, biggest, sizes, sections)
    xlsx.close()
    return bom

def load_boms(bom_folder: Path,
              resources: dict[str, Resource]) -> Boms:
    """load all BOM sheets"""
    status_msg('Loading BOMs', 1)
    bom_files: list[Path] = find_excel_files_in_dir(bom_folder)
    all_boms: Boms = Boms({})
    for bom_file in bom_files:
        bom = load_bom(bom_file, resources)
        all_boms.boms[bom.name] = bom
    return all_boms

if __name__ == "__main__":
    pass
