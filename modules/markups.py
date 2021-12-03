#!/usr/bin/env python
# vim expandtab shiftwidth=2 softtabstop=2
"""
Load model data from sheet

Pass in master_file return data structure
"""
from dataclasses import dataclass
from dataclasses_json import DataClassJsonMixin
from pathlib import Path
from openpyxl import load_workbook # pylint: disable=import-error
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from .utilities import status_msg

@dataclass
class MarkUp(DataClassJsonMixin):
    """Mark-up rates by deprtment entry"""
    markup_1: float
    markup_2: float
    discount: float

@dataclass
class MarkUps(DataClassJsonMixin):
    """Mark-up rates by deprtment dictonary"""
    mark_ups: dict[str, MarkUp]

def load_mark_ups(xlsx_file: Path) -> MarkUps:
    """read makrkup file into object """
    status_msg('Loading Mark Ups', 1)
    status_msg(f'  {xlsx_file.name}', 2)

    mark_up: MarkUp
    mark_ups: MarkUps = MarkUps({})

    xlsx: Workbook = load_workbook(xlsx_file.as_posix(), data_only=True)
    sheet: Worksheet = xlsx.active

    for row in sheet.iter_rows(min_row=2, max_col=4):
        policy, discount = [row[0].value, row[3].value]
        markup_1, markup_2 = [row[1].value, row[2].value]
        if not isinstance(policy, str):
            continue
        mark_up = MarkUp(float(markup_1), float(markup_2), float(discount))
        mark_ups.mark_ups[policy] = mark_up
        status_msg(f"    {policy:22.22} {markup_1:.2f} / {markup_2:.2f}  "
                   f"{discount * 100:5.2f}%", 3)
    xlsx.close()
    return mark_ups

if __name__ == "__main__":
    pass
