#!/usr/bin/env python
# vim expandtab shiftwidth=2 softtabstop=2
"""
Load model data from sheet

Pass in master_file return data structure
"""
from dataclasses import dataclass
from pathlib import Path
from dataclasses_json import dataclass_json
from openpyxl  import load_workbook # pylint: disable=import-error
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from .utilities import status_msg

@dataclass_json
@dataclass
class Model:
    """Information on which sheets make a boat costing sheet"""
    sheet1: str
    sheet2: str
    folder: str

@dataclass_json
@dataclass
class Models:
    """Dictionary of all models"""
    models: dict[str, Model]

def load_models(xlsx_file: Path) -> Models:
    """Build master list of sheets to combine to create costing sheets"""
    status_msg('Loading Boat Models', 1)
    xlsx: Workbook = load_workbook(
        xlsx_file.as_posix(), data_only=True)
    sheet: Worksheet = xlsx.active
    all_models: Models = Models({})
    for row in sheet.iter_rows(min_row=2, max_col=3):
        if not isinstance(row[0].value, str):
            continue
        model: Model = Model(
            row[0].value,
            row[1].value,
            row[2].value)
        all_models.models[model.sheet1] = model
        status_msg(f"    {model}", 3)
    xlsx.close()
    return all_models

if __name__ == "__main__":
    pass
